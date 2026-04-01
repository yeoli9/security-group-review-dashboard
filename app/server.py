"""Flask server for Security Group Review Dashboard."""

__version__ = "1.0.0"

import configparser
import json
import os
import sys
from pathlib import Path
from flask import Flask, jsonify, request, send_from_directory
from flask_cors import CORS

from collector import collect_all
from analyzer import analyze
from governance import load_config as load_governance_config

_APP_DIR = Path(__file__).parent
_PROJECT_DIR = _APP_DIR.parent

app = Flask(__name__, static_folder=str(_APP_DIR / "static"))
app.json.ensure_ascii = False
CORS(app)

# Multi-account cache: { account_id: { data, findings } }
_accounts = {}
CACHE_FILE = str(_PROJECT_DIR / "sg_data_cache.json")


@app.route("/")
def index():
    return send_from_directory("static", "index.html")


@app.route("/api/profiles")
def api_profiles():
    """Read AWS profiles from ~/.aws/config and ~/.aws/credentials."""
    profiles = set()
    for filename in ["config", "credentials"]:
        path = Path.home() / ".aws" / filename
        if not path.exists():
            continue
        cp = configparser.ConfigParser()
        cp.read(str(path))
        for section in cp.sections():
            # config uses "profile xxx", credentials uses "xxx" directly
            name = section.replace("profile ", "") if section.startswith("profile ") else section
            profiles.add(name)
    # default is always available
    profiles.add("default")
    return jsonify(sorted(profiles))


def _resolve_profiles(profiles, region=None):
    """Resolve profiles to account_id + role ARN via STS, then deduplicate.

    For same account_id, prefer profile with 'AdministratorAccess' in ARN.
    """
    import boto3

    resolved = []
    for profile in profiles:
        profile_key = profile or "default"
        try:
            kwargs = {}
            if profile:
                kwargs["profile_name"] = profile
            if region:
                kwargs["region_name"] = region
            session = boto3.Session(**kwargs)
            identity = session.client("sts").get_caller_identity()
            resolved.append({
                "profile": profile,
                "profile_key": profile_key,
                "account_id": identity["Account"],
                "arn": identity.get("Arn", ""),
            })
        except Exception as e:
            resolved.append({
                "profile": profile,
                "profile_key": profile_key,
                "account_id": None,
                "arn": "",
                "error": str(e),
            })

    # Group by account_id, pick best profile per account
    from collections import defaultdict
    by_account = defaultdict(list)
    errors = [r for r in resolved if r["account_id"] is None]
    for r in resolved:
        if r["account_id"]:
            by_account[r["account_id"]].append(r)

    selected = []
    skipped = []
    for acct_id, group in by_account.items():
        if len(group) == 1:
            selected.append(group[0])
        else:
            # Prefer AdministratorAccess
            admin = [r for r in group if "AdministratorAccess" in r["arn"]]
            if admin:
                selected.append(admin[0])
                skipped.extend([r for r in group if r != admin[0]])
            else:
                # No admin found, pick first
                selected.append(group[0])
                skipped.extend(group[1:])

    return selected, skipped, errors


@app.route("/api/collect", methods=["POST"])
def api_collect():
    """Collect fresh data from AWS. Supports multiple profiles."""
    body = request.get_json(silent=True) or {}
    profiles = body.get("profiles", [])
    region = body.get("region")
    dedup = body.get("dedup", True)  # deduplicate same account_id

    # Backward compat: single profile
    if not profiles and body.get("profile"):
        profiles = [body["profile"]]
    if not profiles:
        profiles = [None]  # default profile

    # Resolve and deduplicate
    skipped_info = []
    error_info = []
    if dedup and len(profiles) > 1:
        selected, skipped, errors = _resolve_profiles(profiles, region)
        skipped_info = [{"profile": s["profile_key"], "account_id": s["account_id"], "reason": "duplicate account — kept AdministratorAccess profile"} for s in skipped]
        error_info = [{"profile": e["profile_key"], "status": "error", "message": e.get("error", "auth failed")} for e in errors]
        profiles = [s["profile"] for s in selected]

    results = []
    for profile in profiles:
        profile_key = profile or "default"
        try:
            data = collect_all(profile_name=profile, region_name=region)
            gov_config = load_governance_config()
            findings = analyze(data, governance_config=gov_config)
            data["profile"] = profile_key
            _accounts[profile_key] = {"data": data, "findings": findings}
            all_errors = data.get("collection_errors", [])
            collection_errors = [e for e in all_errors if not e.get("skipped")]
            collection_skipped = [e["service"] for e in all_errors if e.get("skipped")]
            results.append({
                "account_id": data["account_id"],
                "profile": profile_key,
                "status": "ok",
                "summary": findings["summary"],
                "collection_errors": collection_errors,
                "skipped_services": collection_skipped,
            })
        except Exception as e:
            results.append({"profile": profile_key, "status": "error", "message": str(e)})

    results.extend(error_info)

    # Save cache
    _save_cache()

    return jsonify({"status": "ok", "results": results, "skipped": skipped_info})


@app.route("/api/accounts")
def api_accounts():
    """List collected profiles."""
    if not _accounts:
        _load_cache()
    accounts = []
    for profile_key, acct in _accounts.items():
        d = acct["data"]
        f = acct["findings"]
        accounts.append({
            "profile": profile_key,
            "account_id": d.get("account_id", ""),
            "region": d["region"],
            "collection_time": d.get("collection_time", ""),
            "total_sgs": f["summary"]["total_sgs"],
        })
    return jsonify(accounts)


@app.route("/api/data")
def api_data():
    """Return collected data. Filter by account."""
    if not _accounts:
        _load_cache()
    acct_filter = request.args.get("profile")
    if acct_filter and acct_filter in _accounts:
        return jsonify(_accounts[acct_filter]["data"])
    # Return merged or first account
    if not _accounts:
        return jsonify({"status": "error", "message": "No data. Run /api/collect first."}), 404
    merged = _merge_data(acct_filter)
    return jsonify(merged)


@app.route("/api/findings")
def api_findings():
    """Return analysis findings."""
    if not _accounts:
        _load_cache()
    if not _accounts:
        return jsonify({"status": "error", "message": "No data."}), 404
    acct_filter = request.args.get("profile")
    merged = _merge_findings(acct_filter)
    return jsonify(merged)


@app.route("/api/graph")
def api_graph():
    """Return graph data for visualization with VPC grouping."""
    if not _accounts:
        _load_cache()
    if not _accounts:
        return jsonify({"status": "error", "message": "No data."}), 404

    acct_filter = request.args.get("profile")
    vpc_filter = request.args.get("vpc")
    hide_unused = request.args.get("hide_unused", "false") == "true"
    search = request.args.get("search", "").lower()
    risk_filter = request.args.get("risk")  # critical, high, medium, low, unused
    tag_group = request.args.get("tag_group")  # Tag name to color/group by

    # Gather SGs from selected accounts
    all_sgs = []
    all_vpcs = {}
    for profile_key, acct in _accounts.items():
        if acct_filter and profile_key != acct_filter:
            continue
        d = acct["data"]
        all_vpcs.update(d.get("vpcs", {}))
        for sg in d["security_groups"]:
            sg_copy = dict(sg)
            sg_copy["_profile"] = profile_key
            all_sgs.append(sg_copy)

    nodes = []
    edges = []
    vpc_ids_used = set()
    sg_ids_in_graph = set()

    for sg in all_sgs:
        if vpc_filter and sg["vpc_id"] != vpc_filter:
            continue
        if hide_unused and not sg["is_used"]:
            continue
        if search and search not in sg["id"].lower() and search not in sg["name"].lower():
            continue

        node_class = sg.get("risk_level", "low")
        if not sg["is_used"]:
            node_class = "unused"

        # Risk filter
        if risk_filter:
            if risk_filter == "used" and node_class == "unused":
                continue
            elif risk_filter == "high" and node_class not in ("critical", "high"):
                continue
            elif risk_filter == "circular" and not sg.get("in_circular_ref", False):
                continue
            elif risk_filter not in ("used", "high", "circular") and risk_filter != node_class:
                continue

        sg_ids_in_graph.add(sg["id"])
        vpc_ids_used.add(sg["vpc_id"])

        nodes.append({
            "data": {
                "id": sg["id"],
                "label": sg["name"] or sg["id"],
                "type": "sg",
                "nodeClass": node_class,
                "resourceCount": sg["resource_count"],
                "inboundCount": len(sg["inbound_rules"]),
                "outboundCount": len(sg["outbound_rules"]),
                "vpcId": sg["vpc_id"],
                "profile": sg.get("_profile", ""),
                "description": sg.get("description", ""),
                "inCircularRef": sg.get("in_circular_ref", False),
                "tagValue": _get_tag_value(sg, tag_group) if tag_group else None,
            }
        })

        # Add resource nodes
        for res in sg["resources"]:
            res_id = f"{res['type']}:{res['id']}"
            nodes.append({
                "data": {
                    "id": res_id,
                    "label": res.get("name") or res["id"],
                    "type": res["type"].lower(),
                    "nodeClass": res["type"].lower(),
                    "vpcId": sg["vpc_id"],
                }
            })
            edges.append({
                "data": {
                    "id": f"{res_id}->{sg['id']}",
                    "source": res_id,
                    "target": sg["id"],
                    "edgeType": "uses",
                }
            })

        # SG-to-SG reference edges
        for ref_id in sg.get("sg_references", []):
            edges.append({
                "data": {
                    "id": f"{sg['id']}->{ref_id}",
                    "source": sg["id"],
                    "target": ref_id,
                    "edgeType": "sg_ref",
                }
            })

    # Build VPC metadata for client-side boundary drawing
    vpc_meta = {}
    for vpc_id in vpc_ids_used:
        vpc_info = all_vpcs.get(vpc_id, {})
        vpc_name = vpc_info.get("name", "")
        short_id = vpc_id.replace("vpc-", "")[:8]
        vpc_meta[vpc_id] = f"{vpc_name} ({short_id})" if vpc_name else vpc_id

    # Deduplicate resource nodes
    seen_ids = set()
    unique_nodes = []
    for node in nodes:
        nid = node["data"]["id"]
        if nid not in seen_ids:
            seen_ids.add(nid)
            unique_nodes.append(node)

    # Only include edges where both source and target exist
    valid_edges = [e for e in edges if e["data"]["source"] in seen_ids and e["data"]["target"] in seen_ids]

    # Build stats from the filtered SG set (not from findings)
    sg_nodes = [n for n in unique_nodes if n["data"].get("type") == "sg"]
    risky_sgs = [n for n in sg_nodes if n["data"].get("nodeClass") in ("high", "critical")]
    unused_sgs = [n for n in sg_nodes if n["data"].get("nodeClass") == "unused"]
    circular_sgs = [n for n in sg_nodes if n["data"].get("inCircularRef")]
    stats = {
        "total_sgs": len(sg_nodes),
        "used_sgs": len(sg_nodes) - len(unused_sgs),
        "unused_sgs": len(unused_sgs),
        "risky_sgs": len(risky_sgs),
        "circular_sgs": len(circular_sgs),
    }

    # Count default SG warnings from filtered accounts
    default_warnings = []
    for profile_key, acct in _accounts.items():
        if acct_filter and profile_key != acct_filter:
            continue
        default_warnings.extend(acct["findings"].get("default_sg_warnings", []))
    stats["default_sg_warnings"] = len(default_warnings)

    # Count governance warnings from filtered accounts
    gov_warnings = []
    for profile_key, acct in _accounts.items():
        if acct_filter and profile_key != acct_filter:
            continue
        gov_warnings.extend(acct["findings"].get("governance_warnings", []))
    stats["governance_warnings"] = len(gov_warnings)

    return jsonify({"nodes": unique_nodes, "edges": valid_edges, "vpcs": vpc_meta, "stats": stats})


def _get_tag_value(sg, tag_key):
    """Get a tag value from SG by AWS tag key name."""
    for t in sg.get("tags", []):
        if t.get("Key") == tag_key:
            return t["Value"]
    return None


@app.route("/api/tag-values")
def api_tag_values():
    """Return distinct tag keys and their values across all SGs."""
    if not _accounts:
        _load_cache()
    tag_map = {}  # key -> set of values
    for acct in _accounts.values():
        for sg in acct["data"]["security_groups"]:
            for t in sg.get("tags", []):
                key = t.get("Key", "")
                if key:
                    tag_map.setdefault(key, set()).add(t.get("Value", ""))
    return jsonify({k: sorted(v) for k, v in sorted(tag_map.items())})


@app.route("/api/sg/<sg_id>")
def api_sg_detail(sg_id):
    """Return detail for a specific SG."""
    if not _accounts:
        _load_cache()
    for profile_key, acct in _accounts.items():
        for sg in acct["data"]["security_groups"]:
            if sg["id"] == sg_id:
                sg_copy = dict(sg)
                sg_copy["profile"] = profile_key
                sg_copy["account_id"] = acct["data"].get("account_id", "")
                risky = [r for r in acct["findings"].get("risky_rules", []) if r["sg_id"] == sg_id]
                sg_copy["risky_rules"] = risky
                gov = [w for w in acct["findings"].get("governance_warnings", []) if w["sg_id"] == sg_id]
                sg_copy["governance_warnings"] = gov
                return jsonify(sg_copy)
    return jsonify({"status": "not_found"}), 404


@app.route("/api/governance-config")
def api_governance_config():
    """Return resolved governance config."""
    return jsonify(load_governance_config())


@app.route("/api/export/unused")
def api_export_unused():
    """Export unused SGs as JSON/CSV."""
    if not _accounts:
        _load_cache()
    if not _accounts:
        return jsonify([])

    acct_filter = request.args.get("profile")
    fmt = request.args.get("format", "json")

    unused = []
    for profile_key, acct in _accounts.items():
        if acct_filter and profile_key != acct_filter:
            continue
        for u in acct["findings"].get("unused_sgs", []):
            entry = {"profile": profile_key, "account_id": acct["data"].get("account_id", "")}
            entry.update(u)
            unused.append(entry)

    if fmt == "csv":
        import csv, io
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["profile", "account_id", "sg_id", "sg_name", "vpc_id", "description", "inbound_count", "outbound_count"])
        for u in unused:
            writer.writerow([u["profile"], u["account_id"], u["sg_id"], u["sg_name"], u["vpc_id"], u["description"], u["inbound_count"], u["outbound_count"]])
        return output.getvalue(), 200, {"Content-Type": "text/csv", "Content-Disposition": "attachment; filename=unused_sgs.csv"}

    if fmt == "xlsx":
        return _export_xlsx(unused, "unused_sgs")

    return jsonify(unused)


@app.route("/api/export/risky")
def api_export_risky():
    """Export risky SG rules as JSON/CSV/XLSX."""
    if not _accounts:
        _load_cache()
    if not _accounts:
        return jsonify([])

    acct_filter = request.args.get("profile")
    fmt = request.args.get("format", "json")

    risky = []
    for profile_key, acct in _accounts.items():
        if acct_filter and profile_key != acct_filter:
            continue
        for r in acct["findings"].get("risky_rules", []):
            entry = {"profile": profile_key, "account_id": acct["data"].get("account_id", "")}
            entry.update(r)
            risky.append(entry)

    if fmt == "csv":
        import csv, io
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["profile", "account_id", "sg_id", "sg_name", "direction", "protocol", "port", "source", "risk_type", "risk_level", "description"])
        for r in risky:
            writer.writerow([r["profile"], r["account_id"], r["sg_id"], r["sg_name"], r["direction"], r["protocol"], r["port"], r["source"], r["risk_type"], r["risk_level"], r["description"]])
        return output.getvalue(), 200, {"Content-Type": "text/csv", "Content-Disposition": "attachment; filename=risky_rules.csv"}

    if fmt == "xlsx":
        return _export_xlsx(risky, "risky_rules")

    return jsonify(risky)


def _export_xlsx(rows, filename):
    """Export a list of dicts as an XLSX file."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = filename

    if not rows:
        ws.append(["No data"])
    else:
        headers = list(rows[0].keys())
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F6FEB", end_color="1F6FEB", fill_type="solid")
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
        for row in rows:
            ws.append([row.get(h, "") for h in headers])
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue(), 200, {
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "Content-Disposition": f"attachment; filename={filename}.xlsx",
    }


def _merge_data(acct_filter=None):
    """Merge data from all (or filtered) profiles."""
    merged_sgs = []
    merged_vpcs = {}
    for profile_key, acct in _accounts.items():
        if acct_filter and profile_key != acct_filter:
            continue
        d = acct["data"]
        merged_sgs.extend(d["security_groups"])
        merged_vpcs.update(d.get("vpcs", {}))

    first = next(iter(_accounts.values()))["data"]
    return {
        "account_id": acct_filter or "all",
        "region": first["region"],
        "vpcs": merged_vpcs,
        "security_groups": merged_sgs,
        "collection_time": first.get("collection_time", ""),
    }


def _merge_findings(acct_filter=None):
    """Merge findings from all (or filtered) profiles."""
    merged = {"unused_sgs": [], "default_sg_warnings": [], "risky_rules": [], "circular_references": [], "redundant_rules": [], "governance_warnings": [], "summary": {
        "total_sgs": 0, "used_sgs": 0, "unused_sgs": 0,
        "total_inbound_rules": 0, "total_outbound_rules": 0, "vpc_sg_counts": {},
        "default_sg_warnings": 0, "governance_warnings": 0,
    }}
    for profile_key, acct in _accounts.items():
        if acct_filter and profile_key != acct_filter:
            continue
        f = acct["findings"]
        merged["unused_sgs"].extend(f.get("unused_sgs", []))
        merged["default_sg_warnings"].extend(f.get("default_sg_warnings", []))
        merged["risky_rules"].extend(f.get("risky_rules", []))
        merged["circular_references"].extend(f.get("circular_references", []))
        merged["redundant_rules"].extend(f.get("redundant_rules", []))
        merged["governance_warnings"].extend(f.get("governance_warnings", []))
        s = f.get("summary", {})
        merged["summary"]["total_sgs"] += s.get("total_sgs", 0)
        merged["summary"]["used_sgs"] += s.get("used_sgs", 0)
        merged["summary"]["unused_sgs"] += s.get("unused_sgs", 0)
        merged["summary"]["total_inbound_rules"] += s.get("total_inbound_rules", 0)
        merged["summary"]["total_outbound_rules"] += s.get("total_outbound_rules", 0)
        merged["summary"]["default_sg_warnings"] += s.get("default_sg_warnings", 0)
        merged["summary"]["governance_warnings"] += s.get("governance_warnings", 0)
        for vpc_id, cnt in s.get("vpc_sg_counts", {}).items():
            merged["summary"]["vpc_sg_counts"][vpc_id] = merged["summary"]["vpc_sg_counts"].get(vpc_id, 0) + cnt
    return merged


def _save_cache():
    cache = {}
    for acct_id, acct in _accounts.items():
        cache[acct_id] = {"data": acct["data"], "findings": acct["findings"]}
    with open(CACHE_FILE, "w") as f:
        json.dump(cache, f, default=str)


def _load_cache():
    if os.path.exists(CACHE_FILE):
        with open(CACHE_FILE) as f:
            cached = json.load(f)
            # Handle old format (single account)
            if "data" in cached and "findings" in cached:
                acct_id = cached["data"].get("account_id", "unknown")
                _accounts[acct_id] = {"data": cached["data"], "findings": cached["findings"]}
            else:
                # New multi-account format
                for acct_id, acct in cached.items():
                    _accounts[acct_id] = acct


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Security Group Review Dashboard")
    parser.add_argument("port_positional", nargs="?", type=int, default=None,
                        help="Server port (positional, for backward compat)")
    parser.add_argument("--port", "-p", type=int, default=5000, help="Server port (default: 5000)")
    parser.add_argument("--host", default="0.0.0.0", help="Server host (default: 0.0.0.0)")
    parser.add_argument("--cache-file", default=str(_PROJECT_DIR / "sg_data_cache.json"), help="Cache file path")
    parser.add_argument("--no-debug", action="store_true", help="Disable debug mode")
    args = parser.parse_args()

    port = args.port_positional if args.port_positional is not None else args.port
    CACHE_FILE = args.cache_file
    print(f"Starting SG Review Dashboard on http://localhost:{port}")
    app.run(host=args.host, port=port, debug=not args.no_debug)
